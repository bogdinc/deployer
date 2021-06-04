#!/usr/bin/python3
import solcx
import sys
from web3 import Web3, HTTPProvider
import logging
from openpyxl import load_workbook
import re
import os
import json
from openpyxl.utils import get_column_letter


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class Config(object):
    def __init__(self, initial_data):
        for key in initial_data:
            setattr(self, key, initial_data[key])

def load_contracts(filename):
    if not os.path.isfile(filename):
        return {"status": "error", "result": "XLS file with contracts not found."}
    try:
        wb = load_workbook(filename=filename)
    except:
        return {"status": "error", "result": "Invalid XLS file format."}
    ws = wb.active
    ws.cell(row=1, column=6, value='status')
    ws.cell(row=1, column=7, value='address')
    ws.cell(row=1, column=8, value='tx hash')
    wb.save(filename=filename)
    result = list()
    for idx, row in enumerate(ws.iter_rows(min_row=2)):
        try:
            owner = Web3.toChecksumAddress(row[0].value)
            total = int(row[1].value)
            decimals = int(row[2].value)
        except:
            ws.cell(row=idx+2, column=6, value='Invalid')
            continue
        if row[3].value and row[4].value:
            result.append({
                           'owner': owner,
                           'totalSupply': total,
                           'decimals': decimals,
                           'name': row[3].value,
                           'title': row[3].value.replace(" ", ""),
                           'symbol': row[4].value,
                           'row': idx+2
                        })

        else:
            ws.cell(row=idx+2, column=6, value='Invalid')
            continue

    return {"status": "success", "result": result}



def save_contract(filename, data):
    wb = load_workbook(filename=filename)
    ws = wb.active
    if data['status'] == 0:
        ws.cell(row=data['row'], column=6, value='Failed')
    else:
        ws.cell(row=data['row'], column=6, value='Success')
        ws.cell(row=data['row'], column=7, value=data['address'])
    ws.cell(row=data['row'], column=8, value=data['tx_hash'])
    wb.save(filename=filename)


def deploy(abi, bytecode):
    nonce = w3.eth.getTransactionCount(senderAccount.address, "pending")
    contract = w3.eth.contract(abi=abi, bytecode=bytecode)
    transaction = {
        'chainId': chain_id,
        'nonce': nonce,
        'gas': config.GAS_LIMIT,
        'gasPrice': w3.toWei(config.GAS_PRICE, 'gwei')

    }

    txn = contract.constructor().buildTransaction(transaction)
    signed = w3.eth.account.signTransaction(txn, senderAccount.privateKey)
    tx_hash = w3.eth.sendRawTransaction(signed.rawTransaction)
    logger.info("Waiting for receipt...")
    tx_receipt = w3.eth.waitForTransactionReceipt(tx_hash)
    return {"tx_hash": tx_hash.hex(), "status": tx_receipt['status'], "address": tx_receipt.contractAddress}

if __name__ == "__main__":
    with open('config.json', encoding='utf-8') as json_file:
        try:
            data = json.load(json_file)
        except:
            logger.info("Invalid config file. Please check file format. Exiting")
            sys.exit()

    config = Config(data)

    if not os.path.isdir("abis"):
        os.mkdir("abis")

    if not os.path.isfile("base.sol"):
        logger.info("Base source file not found. Exiting...")
        sys.exit()

    with open("base.sol") as f:
        base_source = f.read()

    data = load_contracts(config.INPUT_FILE)

    if data['status'] == "error":
        logger.info(data['result'])
        sys.exit()

    contracts = data['result']

    if len(contracts) == 0:
        logger.info("No valid data for contracts found in XLS file")
        sys.exit()

    if config.TESTMODE:
        node_address = config.NODES['test']
        chain_id = 97
    else:
        node_address = config.NODES['main']
        chain_id = 56

    w3 = Web3(HTTPProvider(node_address))

    if not w3.isConnected():
        logger.info("Failed to connect to node %s. Exiting..." % (node_address))
        sys.exit()

    try:
        senderAccount = w3.eth.account.privateKeyToAccount(config.PRIVATE_KEY)
    except:
        logger.info("Invalid sender private key provided. Exiting...")
        sys.exit()

    solc_folder = solcx.get_solcx_install_folder(solcx_binary_path=None)
    if not os.path.isfile(os.path.join(solc_folder, 'solc-v0.6.12')):
        logger.info("Installing solc...")
        solcx.install_solc('0.6.12')
    solcx.set_solc_version('0.6.12')

    logger.info("Proccessing %s contracts..." % (len(contracts)))
    while len(contracts) > 0:
        current = contracts.pop(0)
        logger.info("Deploying contract %s" % (current['name']))
        source = base_source.replace("[[owner]]", current['owner']).replace("[[symbol]]", '"'+current['symbol']+'"').replace("[[name]]", '"'+current['name']+'"').replace("[[decimals]]", str(current['decimals'])).replace("[[totalSupply]]", str(current['totalSupply'])).replace("[[title]]", current['title'])
        a = solcx.compile_source(
            source,
            #output_values=["abi", "bin-runtime"],
            solc_version="0.6.12",
            #solc_binary = "/opt/solidity/solc-0.6.12",
            optimize=True,
            optimize_runs=200
        )
        abi = a['<stdin>:%s' % (current['title'])]['abi']
        bytecode = a['<stdin>:%s' % (current['title'])]['bin']
        logger.info("Saving ABI...")
        with open("abis/%s.json" % (current['title']), "w") as f:
            f.write(json.dumps(abi))

        try:
            result = deploy(abi, bytecode)
        except Exception as e:
            logger.info("Error: %s" % (str(e)))

        current['status'] = result['status']
        current['address'] = result['address']
        current['tx_hash'] = result['tx_hash']
        save_contract(config.INPUT_FILE, current)


